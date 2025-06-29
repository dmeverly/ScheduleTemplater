import numpy as np

from Solver import Solver
from helpers import (
    weekdays,
    staffRoster,
    ScheduleBalancer,
    validStaffConstraint
)
import matplotlib.pyplot as pp
import pandas as pd
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl import load_workbook

#change PATHIN if you want to read from a different template using .xlsx
#another option is to use the startingTemplate.csv, which is a rough sketch of the template <- requires that undo the comment out of the import_schedule_from_csv method
PATHIN = 'template.xlsx'   # path to the template file to read from - currently set to refeed the last exported solution!
PATHOUT = 'template.xlsx'
STARTERPATHIN = 'startingTemplate.csv'
WEEKS = 6 # if reading from xlsx this needs to be exactly the same number of weeks as weeks to be copied from xlsx
SHIFTLENGTH = 12 #hours
NUM_SHIFTS = 3  # D1, D2, N
DAYS_PER_WEEK = 7

# class with functions to initialize a blank or partially-filled schedule, assign weekends by rotation, import and export schedule templates
# employees are created with constraints from helpers.py, sorted into pools based on day-night shift preference
class Templater:
    def __init__(self):
        self.employees = self._create_employees()
        self.day_pool, self.night_pool, self.float_pool = self._build_pools(self.employees)
        self.unfilled = next(e for e in self.employees if e.name == "UNFILLED")
        self.david = next(e for e in self.employees if e.name == "David")

    def _create_employees(self):
        return [staff.value for staff in staffRoster]

    def _build_pools(self, employees):
        daypool, nightpool, floatpool = [], [], []

        for emp in employees:
            can_do_nights = True
            can_do_days = True

            # Check for day/night shift restrictions (0 shifts per week)
            for c in emp.getConstraints():
                if emp.name == "UNFILLED":
                    continue
                if c.name == validStaffConstraint.NIGHTSHIFTS_PER_WEEK.value and c.val == 0:
                    can_do_nights = False
                if c.name == validStaffConstraint.DAYSHIFTS_PER_WEEK.value and c.val == 0:
                    can_do_days = False

            if can_do_days and not can_do_nights:
                daypool.append(emp)
            elif can_do_nights and not can_do_days:
                nightpool.append(emp)
            else:
                if emp.name != "UNFILLED":
                    floatpool.append(emp)

        return daypool, nightpool, floatpool

    # fills weekends, even week days are David + Another, odd weeks are emp + UNFILLED
    def fillWeekends(self, schedule):
        weeks, days, slots = schedule.shape
        day_rotation_base = [e for e in (self.day_pool + self.float_pool) if e.name != "David" and e.name != "UNFILLED"]
        night_rotation_base = [e for e in (self.float_pool + self.night_pool) if e.name != "David" and e.name != "UNFILLED"]
        
        # extend pools
        day_rotation = day_rotation_base[:]  
        while len(day_rotation) < weeks * 2:
            day_rotation.extend(day_rotation_base[:])

        night_rotation = night_rotation_base[:]  
        while len(night_rotation) < weeks * 2:
            night_rotation.extend(night_rotation_base[:])

        if not day_rotation: # Fallback if base pool is truly empty
            print("Warning: Day rotation pool is critically empty for weekends. Using UNFILLED.")
            day_rotation = [self.unfilled] * (schedule.shape[0] * 2) 
        if not night_rotation: # Fallback if base pool is truly empty
            print("Warning: Night rotation pool is critically empty for weekends. Using UNFILLED.")
            night_rotation = [self.unfilled] * (schedule.shape[0] * 2)

        for w in range(weeks):
            for d in (weekdays.Saturday.value, weekdays.Sunday.value):
                # Sunday inherits from Saturday for consistency in weekend pair
                if d == weekdays.Sunday.value:
                    schedule[w,d,0] = schedule[w,d-1,0]
                    schedule[w,d,1] = schedule[w,d-1,1] 
                    schedule[w,d,2] = schedule[w,d-1,2]
                else: # Saturday
                    if night_rotation:
                        emp_n = night_rotation[0]
                        night_rotation.pop(0)
                        schedule[w,d,2] = emp_n
                        night_rotation.append(emp_n)
                    else:
                        print("night pool empty")
                        exit(1)
                        schedule[w,d,2] = self.unfilled

                    schedule[w,d,1] = self.unfilled 

                    if w % 2 == 0: # Even week: David + one other for day shifts
                        schedule[w,d,0] = self.david # David is D1
                        
                        if day_rotation:
                            other_emp_d2 = day_rotation[0]
                            day_rotation.pop(0)
                            schedule[w,d,1] = other_emp_d2 # Other person is D2
                            day_rotation.append(other_emp_d2)
                        else:
                            print("day pool empty")
                            exit(1)
                            schedule[w,d,1] = self.unfilled # Fallback if day pool empty

                    else: # Odd week: One other person + UNFILLED for day shifts
                        if day_rotation:
                            other_emp_d1 = day_rotation[0]
                            day_rotation.pop(0)
                            schedule[w,d,0] = other_emp_d1 # Other person is D1
                            day_rotation.append(other_emp_d1)
                        else:
                            print("day pool empty")
                            exit(2)
                            schedule[w,d,0] = self.unfilled # Fallback if day pool empty
                        
        return schedule

    #takes number of desired weeks, creates
    def makeTemplate(self, numWeeks: int, fill=False) -> np.ndarray:
        weeks, days, slots = numWeeks, 7, 3
        schedule = np.empty((weeks, days, slots), dtype=object)

        if fill:
            schedule = self.import_schedule_from_csv()
            return schedule

        for week in range(weeks):
            for day in range(days):
                #skip weekends, they are handled separately
                if day in [weekdays.Saturday.value, weekdays.Sunday.value]:
                    continue
                #unfill all weekdays
                for slot in range(3):
                    schedule[week,day,slot] = self.unfilled
                #every other week, assign david to d2 on wednesday
                if week % 2 == 0 and day in [weekdays.Wednesday.value]:
                     schedule[week,day,1] = self.david

        #fill the weekends
        schedule = self.fillWeekends(schedule)

        return schedule

    # import schedule from .xlsx on PATHIN
    # return state as numpy array
    # assumes the format of cells is exactly the same as the export function
    def import_schedule_from_xlsx(self, fill_weekends=False) -> np.ndarray:
        W, D, S = WEEKS, 7, 3
        schedule = np.empty((W, D, S), dtype=object)
        name_map = {e.name: e for e in self.employees}
        name_map[''] = self.unfilled 

        wb = load_workbook(PATHIN, data_only=True)
        ws = wb['Master']

        shift_blocks = {}
        for row in ws.iter_rows(min_row=1, max_col=1):
            cell = row[0]
            if cell.value in ('Day 1 Shifts', 'Day 2 Shifts', 'Night Shifts'):
                label = cell.value.split()[0] 
                if label == 'Day': 
                    full = cell.value 
                    shift_blocks[full.split()[1]] = cell.row  
                else:
                    shift_blocks['N'] = cell.row 

        for shift_idx, key in enumerate(('1','2','N')):
            start = shift_blocks[key]
            data_start = start + 2

            for w in range(W):
                row_idx = data_start + w
                for d in range(D):
                    cell = ws.cell(row=row_idx, column=d+2)
                    name = cell.value if cell.value is not None else ''
                    emp = name_map.get(name, self.unfilled)
                    schedule[w, d, shift_idx] = emp
        
        if fill_weekends:
            schedule = self.fillWeekends(schedule)

        return schedule

    # export schedule from state to PATHOUT
    def export_schedule_to_xlsx(self, schedule: np.ndarray):
        W, _, _ = schedule.shape
        SHIFT_HOURS = 12
        day_names = ['Mo','Tu','We','Th','Fr','Sa','Su']

        tables = {}
        for idx, label in enumerate(('D1','D2','N')):
            data = []
            for w in range(W):
                row = []
                for d in range(7):
                    row.append(schedule[w,d,idx].name if schedule[w,d,idx].name!='UNFILLED' else '')
                data.append(row)

            while len(data) < W:
                data.append(['']*7)

            tables[label] = pd.DataFrame(data, columns=day_names).assign(Week=lambda df: df.index + 1).set_index('Week')

        all_emps = {e for e in schedule.flatten() if e.name!='UNFILLED'}
        summary_rows = []
        for emp in sorted(all_emps, key=lambda e: e.name):
            total_hours = day_shifts = night_shifts = weekday_days = weekend_days = 0
            for w in range(W):
                for d in range(7):
                    worked = False
                    for s in range(3):
                        if schedule[w,d,s] is emp:
                            worked = True
                            total_hours += SHIFT_HOURS
                            if s in (0,1):
                                day_shifts += 1
                            else:
                                night_shifts += 1
                    if worked:
                        if d in (weekdays.Saturday.value, weekdays.Sunday.value):
                            weekend_days += 1
                        else:
                            weekday_days += 1
            summary_rows.append([
                emp.name, total_hours, day_shifts, night_shifts, weekday_days, weekend_days
            ])
        summary_df = pd.DataFrame(
            summary_rows,
            columns=[
                'Employee',
                'Total Hours',
                'Day Shifts',
                'Night Shifts',
                'Weekdays Worked',
                'Weekend Days Worked',
            ]
        )

        personal_dfs = {}
        for emp in sorted(all_emps, key=lambda e: e.name):
            rows = []
            for w in range(W):
                week_hours = 0
                week_row = []
                for d in range(7):
                    slot = ''
                    for s,label in [(0,'D1'),(1,'D2'),(2,'N')]:
                        if schedule[w,d,s] is emp:
                            slot = label
                            week_hours += SHIFT_HOURS
                            break
                    week_row.append(slot)
                rows.append([w+1] + week_row + [week_hours])
            personal_dfs[emp.name] = pd.DataFrame(
                rows,
                columns=['Week'] + day_names + ['Hours']
            )

        with pd.ExcelWriter(PATHOUT, engine='openpyxl') as writer:
            ws = None
            shift_labels = [('D1', 'Day 1'), ('D2', 'Day 2'), ('N', 'Night')]

            
            current_row_offset = 0
            for i, (key, label) in enumerate(shift_labels):
                start_row = current_row_offset + 2
                df = tables[key]

                df.to_excel(
                    writer,
                    sheet_name='Master',
                    startrow=start_row,
                    startcol=0,
                    index=True,  
                    header=True  
                )

                if ws is None:
                    ws = writer.book['Master']

                ws.cell(row=start_row, column=1).value = f"{label} Shifts"
                ws.merge_cells(
                    start_row=start_row,
                    start_column=1,
                    end_row=start_row,
                    end_column=1 + len(day_names) + 1  
                )

                current_row_offset = start_row + len(df) + 3 

            summary_start_row = current_row_offset + 2 

            ws.merge_cells(
                start_row=summary_start_row,
                start_column=1,
                end_row=summary_start_row,
                end_column=len(summary_df.columns)
            )
            ws.cell(row=summary_start_row, column=1).value = "Summary"

            summary_df.to_excel(
                writer,
                sheet_name='Master',
                startrow=summary_start_row + 1,  
                startcol=0,
                index=False, 
                header=True  
            )

            for name, df in personal_dfs.items():
                df.to_excel(writer, sheet_name=name[:31], index=False)

            if ws: 
                for column_cells in ws.columns:
                    max_length = 0
                    column = column_cells[0].column 
                    for cell in column_cells:
                        try:
                            if cell.value is not None:
                                max_length = max(max_length, len(str(cell.value)))
                        except Exception:
                            pass 
                    adjusted_width = (max_length + 2) 
                    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    #this method used specifically for importing a rough sketch of the template, which contained integers to represent employees
    #integer representation compared to employeeMap, which was coded with known length of each employee pool
    #if needed in the future, carefully inspect the length in each pool and extend the employee map as needed
    #read from STARTERPATHIN
    #return state as numpy array
    def import_schedule_from_csv(self) -> np.ndarray:
        
        df = pd.read_csv(STARTERPATHIN, header=None)

        daypool = [e for e in self.day_pool if e.name != "David"]

        #this is hard coded map representation, change as needed for future implementation
        employeeMap = {
            0: self.unfilled,
            1: self.night_pool[0],
            2: self.night_pool[1],
            3: self.float_pool[0],
            4: self.float_pool[1],
            5: daypool[0],
            6: daypool[1],
            7: self.david
        }

        num_weeks = df.shape[0] // NUM_SHIFTS
        # num_days = df.shape[1]  
        schedule = []

        for week in range(num_weeks):
            week_schedule = []
            start_row = week * NUM_SHIFTS
            
            for day in range(DAYS_PER_WEEK):
                day_shift = []
                
                for shift in range(NUM_SHIFTS):
                    value = df.iloc[start_row + shift, day]
                    if pd.isna(value):
                        day_shift.append(self.unfilled)
                    else:
                        day_shift.append(employeeMap.get(int(value), None))
                week_schedule.append(day_shift)
            
            schedule.append(week_schedule)
        schedule = np.array(schedule)

        return schedule

#graph of score over time
def createFigure(epochs, scores):
    _, axis = pp.subplots()
    axis.plot(epochs, scores)
    pp.xlabel("Epochs")
    pp.ylabel("Score")
    pp.show()

#feasibility check to quick fail an unsolvable problem
def isFeasible(employees, total_weeks=WEEKS):
    num_even_weeks = total_weeks // 2
    num_odd_weeks = total_weeks - num_even_weeks
    total_weekday_shifts = num_even_weeks * (3 * 5) + num_odd_weeks * (2 * 5)

    total_weekend_shifts = total_weeks * (2 * 2) 

    total_required_shifts = total_weekday_shifts + total_weekend_shifts
    required_hours = total_required_shifts * SHIFTLENGTH

    total_available_hours = sum(
        next(c.val for c in emp.getConstraints() if c.name == validStaffConstraint.HOURS_PER_PAY_PERIOD.value) * (total_weeks // 2)
        for emp in employees if emp.name != "UNFILLED"
    )

    print(f"Total required shifts over {total_weeks} weeks: {total_required_shifts} (Total Hours: {required_hours})")
    print(f"Total available hours from staff: {total_available_hours}")

    if total_available_hours < required_hours:
        print("⚠️ WARNING: Not enough staff-hours to cover all shifts.")
        return False
    else:
        print("✅ Staff-hour capacity seems sufficient.")
        return True
    
# Main execution block
if __name__ == "__main__":
    # ---------------------------------- INITIALIZATIONS -------------------------------------------
    templater = Templater()    
    employees = templater.employees
    daypool = templater.day_pool
    nightpool = templater.night_pool
    floatpool = templater.float_pool
    unfilled = templater.unfilled

    initial_schedule = templater.import_schedule_from_xlsx()  # <------ comment out to use startingTemplate.csv

    #create initial template using startingTemplate.csv
    #initial_schedule = templater.makeTemplate(WEEKS, fill=True)  <----- uncomment to use the startingTemplate.csv
    
    # Initialize ScheduleBalancer with the initial schedule
    schedule_balancer = ScheduleBalancer(initial_schedule, daypool, nightpool, floatpool, unfilled) 
    # ---------------------------------- SOLVING FUNCTIONS -------------------------------------------
    agent = Solver(schedule_balancer, daypool, nightpool, floatpool, unfilled)

    if not isFeasible(employees):
        exit(0)

    schedule, final_score, epochs, scores = agent.stateHandler()
    
    # ---------------------------------- EVAL AND PRINTING FUNCTIONS -------------------------------------------

    print("\n--- Final Best Solution ---")

    #schedule balancer contains the print functions
    schedule_balancer.state = schedule
    print(schedule_balancer)
    schedule_balancer.printViolations()

    #separate abs and rel violation counts from isValidSchedule
    if not schedule_balancer.isValidSchedule():
        print("Invalid Schedule Created")
  
    print(f"Final Score: {final_score}")

    #print, graph, export to csv  
    templater.export_schedule_to_xlsx(schedule)

    #hours count per employee per week
    weeks, days, slots = schedule.shape
    for emp in employees:
        total_hours = 0
        for week in range(weeks):
            for day in range(7):
                for slot in range(3):
                    if schedule[week, day, slot] is emp:
                        total_hours += SHIFTLENGTH
        print(f"{emp.name}: {total_hours} hrs worked total")

    for emp in employees:
        print(f"{emp.name}:")
        for pay_start in range(0, weeks, 2):
            total_hours = 0
            for w in range(pay_start, min(pay_start+2, weeks)):
                for d in range(7):
                    for s in range(3):
                        if schedule[w, d, s] is emp:
                            total_hours += SHIFTLENGTH
            print(f"  Weeks {pay_start}-{pay_start+1}: {total_hours} hrs")
    
    #print figure
    createFigure(epochs, scores)
